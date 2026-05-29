"""
generatore_classifica.py - Genera il file Excel classifica.xlsx con formattazione professionale.
"""

from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from src.models import PunteggioDettaglio
from src.config import OUTPUT_CLASSIFICA, OUTPUT_DIR
from src.utils import timestamp_ora
from src.logger import get_logger

log = get_logger(__name__)

# ── Palette colori ────────────────────────────────────────────────────────────
C_ORO      = "FFD700"
C_ARGENTO  = "C0C0C0"
C_BRONZO   = "CD7F32"
C_HEADER   = "1A2E45"
C_RIGA_ALT = "EEF2F7"
C_BIANCO   = "FFFFFF"
C_VERDE    = "27AE60"
C_TESTO_H  = "FFFFFF"
C_BORDO    = "BDC3C7"


def _bordo_sottile():
    s = Side(style="thin", color=C_BORDO)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(colore: str) -> PatternFill:
    return PatternFill("solid", start_color=colore, fgColor=colore)


def _header_font() -> Font:
    return Font(bold=True, color=C_TESTO_H, size=11)


def _genera_foglio_classifica(
    ws, punteggi: list[PunteggioDettaglio]
) -> None:
    """Popola il foglio Classifica Generale."""

    # ── Titolo ───────────────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    ws["A1"] = "🏆  CLASSIFICA TOTOMONDIALE 2026"
    ws["A1"].font = Font(bold=True, size=16, color=C_TESTO_H)
    ws["A1"].fill = _fill(C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:N2")
    ws["A2"] = f"Aggiornato al: {timestamp_ora()}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Intestazioni ─────────────────────────────────────────────────────────
    headers = [
        "Pos", "Nome", "TOTALE",
        "Esiti", "Ris.Esatti", "Marcatori", "Gironi",
        "Vincitore", "Finalista", "Capocannoniere",
        "Assistman", "MVP", "Portiere", "Giovane"
    ]
    row_h = 4
    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=row_h, column=col_i, value=h)
        cell.font = _header_font()
        cell.fill = _fill(C_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _bordo_sottile()
    ws.row_dimensions[row_h].height = 36

    # ── Dati ─────────────────────────────────────────────────────────────────
    for i, p in enumerate(punteggi):
        riga = row_h + 1 + i
        pos  = i + 1

        valori = [
            pos,
            p.nome_completo,
            p.punti_totali,
            p.pt_esito,
            p.pt_risultato_esatto,
            p.pt_marcatore,
            p.pt_gironi,
            p.pt_vincitore,
            p.pt_finalista,
            p.pt_capocannoniere,
            p.pt_assistman,
            p.pt_mvp,
            p.pt_miglior_portiere,
            p.pt_miglior_giovane,
        ]

        # Colore riga per podio
        if pos == 1:
            fill_riga = _fill(C_ORO)
            font_riga = Font(bold=True, size=11, color="5D4300")
        elif pos == 2:
            fill_riga = _fill(C_ARGENTO)
            font_riga = Font(bold=True, size=11, color="333333")
        elif pos == 3:
            fill_riga = _fill(C_BRONZO)
            font_riga = Font(bold=True, size=11, color="FFFFFF")
        elif i % 2 == 0:
            fill_riga = _fill(C_RIGA_ALT)
            font_riga = Font(size=10)
        else:
            fill_riga = _fill(C_BIANCO)
            font_riga = Font(size=10)

        for col_i, val in enumerate(valori, 1):
            cell = ws.cell(row=riga, column=col_i, value=val)
            cell.fill = fill_riga
            cell.font = font_riga
            cell.border = _bordo_sottile()
            cell.alignment = Alignment(
                horizontal="center" if col_i != 2 else "left",
                vertical="center"
            )

        ws.row_dimensions[riga].height = 20

    # ── Larghezze colonne ────────────────────────────────────────────────────
    larghezze = [6, 28, 10, 8, 12, 10, 8, 10, 10, 14, 10, 8, 10, 10]
    for col_i, w in enumerate(larghezze, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    # Freeze header
    ws.freeze_panes = ws.cell(row=row_h + 1, column=1)


def _genera_foglio_statistiche(
    ws, punteggi: list[PunteggioDettaglio]
) -> None:
    """Popola il foglio Statistiche."""

    if not punteggi:
        ws["A1"] = "Nessun dato disponibile"
        return

    ws.merge_cells("A1:C1")
    ws["A1"] = "STATISTICHE TOTOMONDIALE 2026"
    ws["A1"].font = Font(bold=True, size=14, color=C_TESTO_H)
    ws["A1"].fill = _fill(C_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    stats = [
        ("Partecipanti totali",   len(punteggi)),
        ("Media punti",           round(sum(p.punti_totali for p in punteggi) / len(punteggi), 1)),
        ("Punteggio massimo",     max(p.punti_totali for p in punteggi)),
        ("Punteggio minimo",      min(p.punti_totali for p in punteggi)),
        ("", ""),
        ("Record risultati esatti",
            max(punteggi, key=lambda x: x.n_risultati_esatti).nome_completo
            + f" ({max(p.n_risultati_esatti for p in punteggi)})"),
        ("Record marcatori",
            max(punteggi, key=lambda x: x.n_marcatori_corretti).nome_completo
            + f" ({max(p.n_marcatori_corretti for p in punteggi)})"),
        ("Record esiti corretti",
            max(punteggi, key=lambda x: x.n_esiti_corretti).nome_completo
            + f" ({max(p.n_esiti_corretti for p in punteggi)})"),
        ("Record punti gironi",
            max(punteggi, key=lambda x: x.pt_gironi).nome_completo
            + f" ({max(p.pt_gironi for p in punteggi)})"),
    ]

    for r_i, (label, valore) in enumerate(stats, 3):
        ws.cell(row=r_i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=r_i, column=2, value=valore)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35


def _genera_foglio_dettaglio(
    ws, punteggi: list[PunteggioDettaglio]
) -> None:
    """Popola il foglio con dettaglio contatori per partecipante."""

    headers = [
        "Pos", "Nome", "Punti Totali",
        "Esiti Corretti", "Risultati Esatti", "Marcatori",
        "Coppie Esatte", "Coppie Inv.", "Singole Qual.", "File"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _header_font()
        cell.fill = _fill(C_HEADER)
        cell.alignment = Alignment(horizontal="center")

    for i, p in enumerate(punteggi):
        ws.append([
            i + 1,
            p.nome_completo,
            p.punti_totali,
            p.n_esiti_corretti,
            p.n_risultati_esatti,
            p.n_marcatori_corretti,
            p.n_gironi_coppia_esatta,
            p.n_gironi_coppia_invertita,
            p.n_gironi_singola,
            p.file_sorgente,
        ])

    for col_i, w in enumerate([6, 28, 12, 14, 16, 10, 14, 12, 13, 25], 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.freeze_panes = "A2"


def genera_excel_classifica(
    punteggi: list[PunteggioDettaglio],
    percorso: Path = OUTPUT_CLASSIFICA,
) -> None:
    """Genera il file Excel classifica con 3 fogli."""

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Foglio 1: Classifica Generale
    ws_class = wb.active
    ws_class.title = "Classifica"
    _genera_foglio_classifica(ws_class, punteggi)

    # Foglio 2: Statistiche
    ws_stat = wb.create_sheet("Statistiche")
    _genera_foglio_statistiche(ws_stat, punteggi)

    # Foglio 3: Dettaglio
    ws_det = wb.create_sheet("Dettaglio")
    _genera_foglio_dettaglio(ws_det, punteggi)

    wb.save(percorso)
    log.info(f"Classifica Excel salvata: {percorso}")
