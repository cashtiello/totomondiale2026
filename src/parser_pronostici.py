"""
parser_pronostici.py - Parser per i file Excel dei partecipanti.

Struttura reale del foglio TABELLONE:
  Riga 2 (idx 1): NOME       -> col C (idx 2)
  Riga 3 (idx 2): COGNOME    -> col C (idx 2)
  Riga 8 (idx 7): header (DATA, GRUPPO, INCONTRO, ..., ESITO, RIS.ESATTO, MARCATORE)
  Righe 9+ (idx 8+): partite

Colonne (0-indexed):
  0=DATA, 1=GRUPPO, 2=INCONTRO, 5=ESITO, 6=RIS.ESATTO, 7=MARCATORE
  9=nome_girone, 11=1°classif, 14=2°classif
  11=label speciale, 12=valore speciale
"""

import re
from pathlib import Path
from typing import Optional

import openpyxl

from src.models import (
    PronosticoPartecipante, PronosticoPartita,
    PronosticoGirone, PronosticoSpeciale
)
from src.utils import (
    normalizza_stringa, normalizza_squadra,
    normalizza_risultato, normalizza_esito, safe_str
)
from src.logger import get_logger

log = get_logger(__name__)

# ── Costanti struttura foglio ─────────────────────────────────────────────────
ROW_NOME    = 1   # riga 2 (0-indexed)
ROW_COGNOME = 2   # riga 3
ROW_START   = 8   # prima riga dati = riga 9 (0-indexed)

COL_DATA      = 0
COL_GRUPPO    = 1
COL_INCONTRO  = 2
COL_ESITO     = 5
COL_RISULTATO = 6
COL_MARCATORE = 7
COL_GIR_NOME  = 9
COL_GIR_1A    = 11
COL_GIR_2A    = 14
COL_SPEC_LBL  = 11
COL_SPEC_V1   = 12


def _cella(row: tuple, col: int) -> Optional[str]:
    if col >= len(row):
        return None
    val = row[col]
    if val is None:
        return None
    # Se è un oggetto datetime (Excel ha interpretato "2-1" come data) → scarta
    import datetime
    if isinstance(val, (datetime.datetime, datetime.date)):
        return None
    s = str(val).strip()
    # Se è un numero intero puro senza trattino → potrebbe essere seriale Excel → scarta
    if s.isdigit():
        return None
    return s if s else None


def _contiene(testo: Optional[str], pattern: str) -> bool:
    if not testo:
        return False
    return pattern.lower() in normalizza_stringa(testo)


def _parse_nome_cognome(rows: list[tuple]) -> tuple[str, str]:
    nome    = _cella(rows[ROW_NOME],    2) or ""
    cognome = _cella(rows[ROW_COGNOME], 2) or ""
    for placeholder in ("inserire nome", "inserire cognome", "nome", "cognome"):
        if normalizza_stringa(nome) == placeholder:
            nome = ""
        if normalizza_stringa(cognome) == placeholder:
            cognome = ""
    return nome, cognome


def _parse_partite(rows: list[tuple]) -> list[PronosticoPartita]:
    partite: list[PronosticoPartita] = []
    for row in rows[ROW_START:]:
        incontro = _cella(row, COL_INCONTRO)
        if not incontro:
            continue
        # Salta intestazioni e label sezioni speciali
        if any(kw in normalizza_stringa(incontro) for kw in (
            "incontro", "data", "gruppo", "accoppiata",
            "vincitrice", "finalista", "capocannoniere",
            "assistman", "mvp", "portiere", "giovane", "team"
        )):
            continue

        # L'esito può essere stringa "1","X","2" oppure intero 1,2
        esito_raw = _cella(row, COL_ESITO)
        esito = normalizza_esito(esito_raw)

        risultato = normalizza_risultato(_cella(row, COL_RISULTATO))
        marcatore = normalizza_squadra(_cella(row, COL_MARCATORE) or "")

        partite.append(PronosticoPartita(
            incontro=normalizza_squadra(incontro),
            esito=esito,
            risultato_esatto=risultato,
            marcatore=marcatore if marcatore else None,
        ))
    return partite


def _parse_gironi(rows: list[tuple]) -> list[PronosticoGirone]:
    gironi: list[PronosticoGirone] = []
    for row in rows[ROW_START:]:
        gir_nome  = _cella(row, COL_GIR_NOME)
        gir_prima = _cella(row, COL_GIR_1A)
        gir_secon = _cella(row, COL_GIR_2A)

        if gir_nome:
            nm = gir_nome.strip().upper()
            if re.fullmatch(r"[A-L]", nm):
                prima  = normalizza_squadra(gir_prima) if gir_prima else None
                seconda = normalizza_squadra(gir_secon) if gir_secon else None
                if prima or seconda:
                    gironi.append(PronosticoGirone(
                        girone=nm,
                        prima=prima,
                        seconda=seconda,
                    ))
    return gironi


def _parse_speciali(rows: list[tuple]) -> PronosticoSpeciale:
    spec = PronosticoSpeciale()
    i = 0
    while i < len(rows):
        row = rows[i]
        lbl = _cella(row, COL_SPEC_LBL)
        if not lbl:
            i += 1
            continue

        lbl_n = normalizza_stringa(lbl)

        if _contiene(lbl_n, "vincitrice"):
            # Il valore è nella riga SUCCESSIVA, colonna 11
            if i + 1 < len(rows):
                spec.vincitore = normalizza_squadra(_cella(rows[i+1], COL_SPEC_LBL) or "") or None

        elif _contiene(lbl_n, "finalista"):
            # TEAM 1 e TEAM 2 sono nelle righe successive
            for j in range(i + 1, min(i + 5, len(rows))):
                r2 = rows[j]
                lbl2 = _cella(r2, COL_SPEC_LBL)
                val2 = _cella(r2, COL_SPEC_V1)
                if lbl2 and _contiene(lbl2, "team 1"):
                    spec.finalista_1 = normalizza_squadra(val2 or "") or None
                elif lbl2 and _contiene(lbl2, "team 2"):
                    spec.finalista_2 = normalizza_squadra(val2 or "") or None

        elif _contiene(lbl_n, "capocannoniere"):
            if i + 1 < len(rows):
                spec.capocannoniere = normalizza_squadra(_cella(rows[i+1], COL_SPEC_LBL) or "") or None

        elif _contiene(lbl_n, "assistman"):
            if i + 1 < len(rows):
                spec.assistman = normalizza_squadra(_cella(rows[i+1], COL_SPEC_LBL) or "") or None

        elif _contiene(lbl_n, "mvp"):
            if i + 1 < len(rows):
                spec.mvp = normalizza_squadra(_cella(rows[i+1], COL_SPEC_LBL) or "") or None

        elif _contiene(lbl_n, "miglior portiere"):
            if i + 1 < len(rows):
                spec.miglior_portiere = normalizza_squadra(_cella(rows[i+1], COL_SPEC_LBL) or "") or None

        elif _contiene(lbl_n, "miglior giovane") or _contiene(lbl_n, "u21"):
            if i + 1 < len(rows):
                spec.miglior_giovane = normalizza_squadra(_cella(rows[i+1], COL_SPEC_LBL) or "") or None

        i += 1
    return spec


def parse_file_partecipante(percorso: Path) -> Optional[PronosticoPartecipante]:
    try:
        wb = openpyxl.load_workbook(percorso, data_only=True)
    except Exception as e:
        log.error(f"Impossibile aprire {percorso.name}: {e}")
        return None

    foglio = None
    for nome_foglio in wb.sheetnames:
        if "tabellone" in nome_foglio.lower():
            foglio = wb[nome_foglio]
            break
    if foglio is None:
        foglio = wb.active
    if foglio is None:
        log.error(f"{percorso.name}: workbook vuoto")
        return None

    rows = [tuple(row) for row in foglio.iter_rows(values_only=True)]

    if len(rows) < ROW_START + 1:
        log.warning(f"{percorso.name}: foglio troppo corto ({len(rows)} righe)")
        return None

    try:
        nome, cognome = _parse_nome_cognome(rows)
        partite       = _parse_partite(rows)
        gironi        = _parse_gironi(rows)
        speciali      = _parse_speciali(rows)
    except Exception as e:
        log.error(f"{percorso.name}: errore parsing - {e}", exc_info=True)
        return None

    if not nome and not cognome:
        nome = percorso.stem
        log.warning(f"{percorso.name}: nome/cognome vuoti, uso nome file: '{nome}'")

    log.info(
        f"Letto {percorso.name}: {nome} {cognome} | "
        f"{len(partite)} partite | {len(gironi)} gironi | "
        f"vincitore={speciali.vincitore} | capo={speciali.capocannoniere}"
    )

    return PronosticoPartecipante(
        nome=nome,
        cognome=cognome,
        file_sorgente=percorso.name,
        partite=partite,
        gironi=gironi,
        speciali=speciali,
    )