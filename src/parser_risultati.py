"""
parser_risultati.py - Legge il file Excel dei risultati reali.
"""

from pathlib import Path
from typing import Optional
import json

import openpyxl

from src.models import RisultatiReali, RisultatoPartita, RisultatoGirone
from src.utils import (
    normalizza_stringa, normalizza_squadra,
    normalizza_risultato, safe_str
)
from src.config import RISULTATI_FILE
from src.logger import get_logger

log = get_logger(__name__)

RISULTATI_SPECIALI_FILE = Path(__file__).parent.parent / "data" / "risultati_speciali.json"

CHIAVI_SPECIALI = {
    "vincitore":        ("vincitore", "vincitrice", "vincitrice competizione"),
    "finalista_1":      ("finalista 1", "team 1", "finalista1"),
    "finalista_2":      ("finalista 2", "team 2", "finalista2"),
    "capocannoniere":   ("capocannoniere",),
    "assistman":        ("assistman",),
    "mvp":              ("mvp", "mvp torneo"),
    "miglior_portiere": ("miglior portiere",),
    "miglior_giovane":  ("miglior giovane", "u21", "miglior giovane u21"),
}


def _match_chiave(raw: str) -> Optional[str]:
    n = normalizza_stringa(raw)
    for campo, varianti in CHIAVI_SPECIALI.items():
        for v in varianti:
            if v in n:
                return campo
    return None


def _leggi_speciali_da_json() -> dict[str, Optional[str]]:
    """Legge i risultati speciali da data/risultati_speciali.json."""
    speciali = {}
    try:
        if RISULTATI_SPECIALI_FILE.exists():
            with open(RISULTATI_SPECIALI_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            mapping = {
                "vincitore":        "vincitore",
                "finalista_1":      "finalista_1",
                "finalista_2":      "finalista_2",
                "capocannoniere":   "capocannoniere",
                "assistman":        "assistman",
                "mvp":              "mvp",
                "miglior_portiere": "miglior_portiere",
                "miglior_giovane":  "miglior_giovane",
            }
            for key, campo in mapping.items():
                val = data.get(key, "")
                if val:
                    speciali[campo] = val.strip()
                    log.info(f"  📋 Speciale da JSON: {campo} = {val}")
    except Exception as e:
        log.warning(f"Errore lettura risultati_speciali.json: {e}")
    return speciali


def _leggi_foglio_partite(ws) -> dict[str, RisultatoPartita]:
    risultati: dict[str, RisultatoPartita] = {}
    header_trovato = False

    for row in ws.iter_rows(values_only=True):
        vals = [safe_str(v) for v in row]
        if not any(vals):
            continue
        if not header_trovato:
            ns = [normalizza_stringa(v) for v in vals]
            if any("partita" in n or "incontro" in n for n in ns):
                header_trovato = True
            continue

        partita   = normalizza_squadra(vals[0]) if vals[0] else None
        risultato = normalizza_risultato(vals[1]) if len(vals) > 1 else None
        marcatore = normalizza_squadra(vals[2]) if len(vals) > 2 and vals[2] else None

        if not partita:
            continue

        risultati[partita] = RisultatoPartita(
            incontro=partita,
            risultato=risultato,
            marcatore=marcatore or None,
        )

    return risultati


def _leggi_foglio_gironi(ws) -> dict[str, RisultatoGirone]:
    gironi: dict[str, RisultatoGirone] = {}
    header_trovato = False

    for row in ws.iter_rows(values_only=True):
        vals = [safe_str(v) for v in row]
        if not any(vals):
            continue
        if not header_trovato:
            ns = [normalizza_stringa(v) for v in vals]
            if any("girone" in n for n in ns):
                header_trovato = True
            continue

        girone = vals[0].strip().upper() if vals[0] else None
        prima  = normalizza_squadra(vals[1]) if len(vals) > 1 and vals[1] else None
        secon  = normalizza_squadra(vals[2]) if len(vals) > 2 and vals[2] else None

        if not girone:
            continue

        gironi[girone] = RisultatoGirone(
            girone=girone,
            prima=prima or None,
            seconda=secon or None,
        )

    return gironi


def _leggi_foglio_speciali(ws) -> dict[str, Optional[str]]:
    speciali: dict[str, Optional[str]] = {}
    header_trovato = False

    for row in ws.iter_rows(values_only=True):
        vals = [safe_str(v) for v in row]
        if not any(vals):
            continue
        if not header_trovato:
            ns = [normalizza_stringa(v) for v in vals]
            if any("voce" in n or "categoria" in n for n in ns):
                header_trovato = True
            continue

        label = vals[0]
        valore = vals[1] if len(vals) > 1 else None

        campo = _match_chiave(label)
        if campo:
            speciali[campo] = normalizza_squadra(valore) if valore else None

    return speciali


def leggi_risultati(percorso: Path = RISULTATI_FILE) -> RisultatiReali:
    risultati = RisultatiReali()

    if not percorso.exists():
        log.warning(f"File risultati non trovato: {percorso}")
        log.info("Verrà creato un file di esempio. Compilarlo e rieseguire.")
        _crea_file_esempio(percorso)
        return risultati

    try:
        wb = openpyxl.load_workbook(percorso, data_only=True)
    except Exception as e:
        log.error(f"Impossibile aprire {percorso}: {e}")
        return risultati

    sheet_names_lower = {n.lower(): n for n in wb.sheetnames}

    # ── Partite ──────────────────────────────────────────────────────────────
    foglio_partite = None
    for candidato in ("partite", "risultati", "tabellone"):
        if candidato in sheet_names_lower:
            foglio_partite = wb[sheet_names_lower[candidato]]
            break
    if foglio_partite is None:
        foglio_partite = wb.active

    risultati.partite = _leggi_foglio_partite(foglio_partite)
    log.info(f"Risultati partite caricati: {len(risultati.partite)}")

    # ── Gironi ───────────────────────────────────────────────────────────────
    if "gironi" in sheet_names_lower:
        risultati.gironi = _leggi_foglio_gironi(wb[sheet_names_lower["gironi"]])
        log.info(f"Risultati gironi caricati: {len(risultati.gironi)}")

    # ── Speciali: prima JSON, poi Excel come fallback ─────────────────────────
    speciali = _leggi_speciali_da_json()

    # Fallback Excel per campi non presenti nel JSON
    if "speciali" in sheet_names_lower:
        speciali_excel = _leggi_foglio_speciali(wb[sheet_names_lower["speciali"]])
        for campo, valore in speciali_excel.items():
            if campo not in speciali and valore:
                speciali[campo] = valore

    risultati.vincitore         = speciali.get("vincitore")
    risultati.finalista_1       = speciali.get("finalista_1")
    risultati.finalista_2       = speciali.get("finalista_2")
    risultati.capocannoniere    = speciali.get("capocannoniere")
    risultati.assistman         = speciali.get("assistman")
    risultati.mvp               = speciali.get("mvp")
    risultati.miglior_portiere  = speciali.get("miglior_portiere")
    risultati.miglior_giovane   = speciali.get("miglior_giovane")

    log.info(f"Speciali caricati: {speciali}")
    return risultati


def _crea_file_esempio(percorso: Path) -> None:
    percorso.parent.mkdir(parents=True, exist_ok=True)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws_p = wb.active
    ws_p.title = "PARTITE"
    header = ["PARTITA", "RISULTATO", "MARCATORE"]
    ws_p.append(header)
    for cell in ws_p[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="2E4057")
        cell.alignment = Alignment(horizontal="center")
    ws_p.column_dimensions["A"].width = 30
    ws_p.column_dimensions["B"].width = 15
    ws_p.column_dimensions["C"].width = 20

    ws_g = wb.create_sheet("GIRONI")
    ws_g.append(["GIRONE", "1° CLASSIFICATA", "2° CLASSIFICATA"])
    for cell in ws_g[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="2E4057")
        cell.alignment = Alignment(horizontal="center")
    for g in "ABCDEFGHIJKL":
        ws_g.append([g, "", ""])
    ws_g.column_dimensions["A"].width = 10
    ws_g.column_dimensions["B"].width = 25
    ws_g.column_dimensions["C"].width = 25

    ws_s = wb.create_sheet("SPECIALI")
    ws_s.append(["VOCE", "VALORE"])
    for cell in ws_s[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="2E4057")
        cell.alignment = Alignment(horizontal="center")
    for voce in [("VINCITORE",""),("FINALISTA 1",""),("FINALISTA 2",""),
                 ("CAPOCANNONIERE",""),("ASSISTMAN",""),("MVP TORNEO",""),
                 ("MIGLIOR PORTIERE",""),("MIGLIOR GIOVANE U21","")]:
        ws_s.append(voce)
    ws_s.column_dimensions["A"].width = 25
    ws_s.column_dimensions["B"].width = 25

    wb.save(percorso)
    log.info(f"File risultati di esempio creato: {percorso}")
