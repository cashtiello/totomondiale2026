"""
aggiorna_e_calcola.py
Script eseguito da GitHub Actions ogni ora.
"""

import os
import sys
import json
import urllib.request
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.logger import setup_logging, get_logger
from src.config import RISULTATI_FILE

RISULTATI_MANUALI_FILE = Path(__file__).parent.parent / "data" / "risultati_manuali.json"
ELIMINATORIE_DIR = Path(__file__).parent.parent / "data" / "eliminatorie"
RISULTATI_SPECIALI_FILE = Path(__file__).parent.parent / "data" / "risultati_speciali.json"

from src.excel_reader import leggi_tutti_pronostici
from src.parser_risultati import leggi_risultati
from src.calcolo_punti import calcola_tutti_punteggi
from src.generatore_classifica import genera_excel_classifica
from src.html_generator import genera_html

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

setup_logging()
log = get_logger("aggiorna_e_calcola")

API_KEY  = os.environ.get("FOOTBALL_DATA_KEY", "")
API_BASE = "https://worldcup26.ir"

NOMI_SQUADRE = {
    "Mexico": "Messico", "South Africa": "Sud Africa",
    "Korea Republic": "Corea del Sud", "South Korea": "Corea del Sud",
    "Korea DPR": "Corea del Nord", "Czech Republic": "R.Ceca",
    "Czechia": "R.Ceca", "Czech Rep.": "R.Ceca",
    "Canada": "Canada", "Bosnia and Herzegovina": "Bosnia",
    "Bosnia-Herzegovina": "Bosnia", "Bosnia & Herzegovina": "Bosnia",
    "Bosnia": "Bosnia", "Qatar": "Qatar", "Switzerland": "Svizzera",
    "Brazil": "Brasile", "Brasil": "Brasile", "Morocco": "Marocco",
    "Haiti": "Haiti", "Scotland": "Scozia",
    "USA": "USA", "United States": "USA", "United States of America": "USA",
    "Paraguay": "Paraguay", "Australia": "Australia",
    "Turkey": "Turchia", "Türkiye": "Turchia",
    "Germany": "Germania", "Curaçao": "Curacao", "Curacao": "Curacao",
    "Ivory Coast": "Costa d'Avorio", "Côte d'Ivoire": "Costa d'Avorio",
    "Cote d'Ivoire": "Costa d'Avorio", "Ecuador": "Ecuador",
    "Netherlands": "Olanda", "Holland": "Olanda", "Japan": "Giappone",
    "Sweden": "Svezia", "Tunisia": "Tunisia",
    "Belgium": "Belgio", "Egypt": "Egitto", "Iran": "Iran",
    "New Zealand": "Nuova Zelanda",
    "Spain": "Spagna", "Cape Verde": "Capo Verde", "Cabo Verde": "Capo Verde",
    "Saudi Arabia": "Arabia Saudita", "KSA": "Arabia Saudita", "Uruguay": "Uruguay",
    "France": "Francia", "Senegal": "Senegal", "Iraq": "Iraq", "Norway": "Norvegia",
    "Argentina": "Argentina", "Algeria": "Algeria", "Austria": "Austria",
    "Jordan": "Giordania",
    "Portugal": "Portogallo", "Congo": "Congo", "DR Congo": "Congo",
    "Democratic Republic of the Congo": "Congo", "Republic of Congo": "Congo",
    "Uzbekistan": "Uzbekistan", "Colombia": "Colombia",
    "England": "Inghilterra", "Croatia": "Croazia", "Ghana": "Ghana", "Panama": "Panama",
}

def traduci(nome: str) -> str:
    return NOMI_SQUADRE.get(nome, nome)

GIRONI_SQUADRE = {
    'A': ['Messico', 'Sud Africa', 'Corea del Sud', 'Repubblica Ceca'],
    'B': ['Canada', 'Bosnia Erzegovina', 'Qatar', 'Svizzera'],
    'C': ['Brasile', 'Marocco', 'Haiti', 'Scozia'],
    'D': ['USA', 'Paraguay', 'Australia', 'Turchia'],
    'E': ['Germania', 'Curacao', "Costa d'Avorio", 'Ecuador'],
    'F': ['Olanda', 'Giappone', 'Svezia', 'Tunisia'],
    'G': ['Belgio', 'Egitto', 'Iran', 'Nuova Zelanda'],
    'H': ['Spagna', 'Capo Verde', 'Arabia Saudita', 'Uruguay'],
    'I': ['Francia', 'Senegal', 'Iraq', 'Norvegia'],
    'J': ['Argentina', 'Algeria', 'Austria', 'Giordania'],
    'K': ['Portogallo', 'Congo', 'Uzbekistan', 'Colombia'],
    'L': ['Inghilterra', 'Croazia', 'Ghana', 'Panama'],
}

TUTTE_LE_PARTITE = [
    ("Messico-Sud Africa","A"), ("Corea del Sud-R.Ceca","A"),
    ("Canada-Bosnia","B"), ("USA-Paraguay","D"), ("Qatar-Svizzera","B"),
    ("Brasile-Marocco","C"), ("Haiti-Scozia","C"), ("Australia-Turchia","D"),
    ("Germania-Curacao","E"), ("Olanda-Giappone","F"),
    ("Costa d'Avorio-Ecuador","E"), ("Svezia-Tunisia","F"),
    ("Spagna-Capo Verde","H"), ("Belgio-Egitto","G"),
    ("Arabia Saudita-Uruguay","H"), ("Iran-Nuova Zelanda","G"),
    ("Francia-Senegal","I"), ("Iraq-Norvegia","I"),
    ("Argentina-Algeria","J"), ("Austria-Giordania","J"),
    ("Portogallo-Congo","K"), ("Inghilterra-Croazia","L"),
    ("Ghana-Panama","L"), ("Uzbekistan-Colombia","K"),
    ("R.Ceca-Sud Africa","A"), ("Svizzera-Bosnia","B"),
    ("Canada-Qatar","B"), ("Messico-Corea del Sud","A"),
    ("USA-Australia","D"), ("Scozia-Marocco","C"),
    ("Brasile-Haiti","C"), ("Turchia-Paraguay","D"),
    ("Olanda-Svezia","F"), ("Germania-Costa d'Avorio","E"),
    ("Ecuador-Curacao","E"), ("Tunisia-Giappone","F"),
    ("Spagna-Arabia Saudita","H"), ("Belgio-Iran","G"),
    ("Uruguay-Capo Verde","H"), ("Nuova Zelanda-Egitto","G"),
    ("Argentina-Austria","J"), ("Francia-Iraq","I"),
    ("Norvegia-Senegal","I"), ("Giordania-Algeria","J"),
    ("Portogallo-Uzbekistan","K"), ("Inghilterra-Ghana","L"),
    ("Panama-Croazia","L"), ("Colombia-Congo","K"),
    ("Svizzera-Canada","B"), ("Bosnia-Qatar","B"),
    ("Scozia-Brasile","C"), ("Marocco-Haiti","C"),
    ("R.Ceca-Messico","A"), ("Sud Africa-Corea del Sud","A"),
    ("Ecuador-Germania","E"), ("Curacao-Costa d'Avorio","E"),
    ("Tunisia-Olanda","F"), ("Giappone-Svezia","F"),
    ("Paraguay-Australia","D"), ("Turchia-USA","D"),
    ("Norvegia-Francia","I"), ("Senegal-Iraq","I"),
    ("Uruguay-Spagna","H"), ("Capo Verde-Arabia Saudita","H"),
    ("Egitto-Iran","G"), ("Nuova Zelanda-Belgio","G"),
    ("Panama-Inghilterra","L"), ("Croazia-Ghana","L"),
    ("Colombia-Portogallo","K"), ("Congo-Uzbekistan","K"),
    ("Giordania-Argentina","J"), ("Algeria-Austria","J"),
]


def _api(endpoint: str) -> dict:
    url = f"{API_BASE}/{endpoint}"
    req = urllib.request.Request(url, headers={"User-Agent": "TotoMondiale/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode())
    except Exception as e:
        log.error(f"Errore API {endpoint}: {e}")
        return {}


def _parse_scorers(scorers_str: str) -> tuple[list[str], bool]:
    if not scorers_str or scorers_str == "null":
        return [], False
    import re
    scorers_str = scorers_str.strip('{}')
    scorers_str = re.sub(r'[\u0022\u0027\u201c\u201d\u2018\u2019\u00ab\u00bb]', '', scorers_str)
    nomi = []
    ha_autogol = False
    for s in scorers_str.split(","):
        s = s.strip()
        if not s or s == "null":
            continue
        nome = re.sub(r"\s+\d+\+?'?$", "", s).strip()
        nome = re.sub(r"\s+\d+\+?['\u2019]?$", "", nome).strip()
        if "autogoal" in nome.lower() or "own goal" in nome.lower():
            ha_autogol = True
        elif nome:
            nomi.append(nome)
    return nomi, ha_autogol


def scarica_risultati() -> dict[str, dict]:
    log.info("Scarico partite da worldcup26.ir...")
    data = _api("get/games")
    risultati = {}
    for match in data.get("games", []):
        if match.get("finished", "FALSE").upper() != "TRUE":
            continue
        home = traduci(match.get("home_team_name_en", ""))
        away = traduci(match.get("away_team_name_en", ""))
        gh = int(match.get("home_score", 0) or 0)
        ga = int(match.get("away_score", 0) or 0)
        nome = f"{home}-{away}"
        gol_home, auto_home = _parse_scorers(match.get("home_scorers", "null"))
        gol_away, auto_away = _parse_scorers(match.get("away_scorers", "null"))
        gol = gol_home + gol_away
        ha_autogol = auto_home or auto_away
        marcatori_str = ", ".join(gol)
        if ha_autogol:
            marcatori_str = (marcatori_str + ", autogol").strip(", ")
        risultati[nome] = {"risultato": f"{gh}-{ga}", "marcatori": marcatori_str}
        log.info(f"  {nome}: {gh}-{ga} | {marcatori_str}")
    log.info(f"Scaricate {len(risultati)} partite completate")
    return risultati


def scarica_gironi() -> dict[str, dict]:
    log.info("Scarico classifiche gironi...")
    data = _api("get/groups")
    classifiche = {}
    for gruppo in data.get("groups", []):
        lettera = gruppo.get("name", "").replace("Group ", "").strip()
        teams = gruppo.get("teams", [])
        if len(teams) >= 2:
            teams_sorted = sorted(teams, key=lambda t: int(t.get("pts", 0) or 0), reverse=True)
            classifiche[lettera] = {
                "prima":   traduci(teams_sorted[0].get("team_name_en", "")),
                "seconda": traduci(teams_sorted[1].get("team_name_en", "")),
            }
    log.info(f"Classifiche gironi: {len(classifiche)}")
    return classifiche


def aggiorna_json_manuali(partite_api: dict) -> tuple[dict, dict]:
    try:
        if not RISULTATI_MANUALI_FILE.exists():
            log.warning("risultati_manuali.json non trovato")
            return {}, {}

        with open(RISULTATI_MANUALI_FILE, "r", encoding="utf-8") as f:
            json_data = json.load(f)

        partite_json = json_data.get("partite", {})
        dati_manuali = {}

        for partita, dati in partite_json.items():
            is_manuale = dati.get("_manuale", False)
            api_info = partite_api.get(partita, {})
            if is_manuale:
                dati_manuali[partita] = {
                    "risultato": dati.get("risultato", ""),
                    "marcatori": dati.get("marcatori", ""),
                }
                log.info(f"  ✋ MANUALE protetto: {partita}")
            elif api_info:
                partite_json[partita]["risultato"] = api_info.get("risultato", "")
                partite_json[partita]["marcatori"] = api_info.get("marcatori", "")

        gironi_json = json_data.get("gironi", {})
        gironi_manuali = {}
        for lettera, dati in gironi_json.items():
            if dati.get("_manuale", False):
                gironi_manuali[lettera] = {
                    "prima":   dati.get("prima", ""),
                    "seconda": dati.get("seconda", ""),
                }
                log.info(f"  ✋ GIRONE MANUALE {lettera}: {dati.get('prima','')} / {dati.get('seconda','')}")

        json_data["partite"] = partite_json
        with open(RISULTATI_MANUALI_FILE, "w", encoding="utf-8") as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)

        log.info(f"risultati_manuali.json aggiornato ({len(dati_manuali)} partite manuali, {len(gironi_manuali)} gironi manuali)")
        return dati_manuali, gironi_manuali

    except Exception as e:
        log.error(f"Errore aggiornamento JSON manuali: {e}")
        return {}, {}


def carica_eliminatorie() -> dict:
    """
    Carica tutti i file eliminatorie da data/eliminatorie/.
    Struttura: {fase: {partecipante: {partite: [...]}}}
    """
    eliminatorie = {}
    if not ELIMINATORIE_DIR.exists():
        return eliminatorie

    for fase_file in sorted(ELIMINATORIE_DIR.glob("*.json")):
        fase = fase_file.stem  # es. "sedicesimi"
        try:
            with open(fase_file, encoding="utf-8") as f:
                data = json.load(f)
            eliminatorie[fase] = data
            log.info(f"Caricata fase {fase}: {len(data)} partecipanti")
        except Exception as e:
            log.error(f"Errore lettura {fase_file}: {e}")

    return eliminatorie


def scrivi_risultati(partite: dict, gironi: dict) -> None:
    bordo = lambda: Border(
        left=Side(style='thin', color='BDC3C7'), right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'), bottom=Side(style='thin', color='BDC3C7'),
    )
    h_fill = PatternFill("solid", start_color="2E4057")
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_alg  = Alignment(horizontal="center", vertical="center")
    f_alt  = [PatternFill("solid", start_color="F8F9FA"), PatternFill("solid", start_color="FFFFFF")]

    dati_manuali, gironi_manuali = aggiorna_json_manuali(partite)

    partite_json_esistente = {}
    try:
        if RISULTATI_MANUALI_FILE.exists():
            with open(RISULTATI_MANUALI_FILE, "r", encoding="utf-8") as f:
                partite_json_esistente = json.load(f).get("partite", {})
    except Exception:
        pass

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "PARTITE"
    for col, h in enumerate(["PARTITA", "RISULTATO", "MARCATORE"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = h_alg; c.border = bordo()
    ws.row_dimensions[1].height = 24

    for i, (incontro, _) in enumerate(TUTTE_LE_PARTITE):
        info = partite.get(incontro, {})
        manuale = dati_manuali.get(incontro, {})
        json_esistente = partite_json_esistente.get(incontro, {})
        if manuale:
            risultato_finale = manuale.get("risultato", "")
            marcatori_finali = manuale.get("marcatori", "")
            log.info(f"  ✋ {incontro}: MANUALE ({risultato_finale} | {marcatori_finali})")
        elif info:
            risultato_finale = info.get("risultato", "")
            marcatori_finali = info.get("marcatori", "")
        else:
            risultato_finale = json_esistente.get("risultato", "")
            marcatori_finali = json_esistente.get("marcatori", "")
            if risultato_finale:
                log.info(f"  {incontro}: API down, uso JSON esistente")
        for col, val in enumerate([incontro, risultato_finale, marcatori_finali], 1):
            c = ws.cell(row=i+2, column=col, value=val)
            c.fill = f_alt[i % 2]
            c.alignment = Alignment(horizontal="left" if col==1 else "center", vertical="center")
            c.border = bordo()
        ws.row_dimensions[i+2].height = 18

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 35

    ws_g = wb.create_sheet("GIRONI")
    for col, h in enumerate(["GIRONE", "1° CLASSIFICATA", "2° CLASSIFICATA"], 1):
        c = ws_g.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = h_alg; c.border = bordo()
    ws_g.row_dimensions[1].height = 24

    gir_fill = PatternFill("solid", start_color="EEF2F7")
    for i, (lettera, squadre) in enumerate(GIRONI_SQUADRE.items(), 2):
        if lettera in gironi_manuali:
            info = gironi_manuali[lettera]
            log.info(f"  ✋ Girone {lettera}: MANUALE ({info.get('prima','')} / {info.get('seconda','')})")
        else:
            info = gironi.get(lettera, {})

        c = ws_g.cell(row=i, column=1, value=lettera)
        c.font = Font(bold=True, size=11, color="1A2E45")
        c.fill = gir_fill; c.alignment = h_alg; c.border = bordo()

        formula = '"' + ','.join(squadre) + '"'
        for col, val in [(2, info.get("prima","")), (3, info.get("seconda",""))]:
            cell = ws_g.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = bordo()
            dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
            ws_g.add_data_validation(dv)
            dv.add(cell)
        ws_g.row_dimensions[i].height = 22

    ws_g.column_dimensions['A'].width = 10
    ws_g.column_dimensions['B'].width = 28
    ws_g.column_dimensions['C'].width = 28

    ws_s = wb.create_sheet("SPECIALI")
    for col, h in enumerate(["VOCE", "VALORE"], 1):
        c = ws_s.cell(row=1, column=col, value=h)
        c.font = h_font; c.fill = h_fill; c.alignment = h_alg; c.border = bordo()

    speciali_esistenti = {}
    # 1. Prima prova a leggere da risultati_speciali.json
    try:
        if RISULTATI_SPECIALI_FILE.exists():
            with open(RISULTATI_SPECIALI_FILE, "r", encoding="utf-8") as f:
                rs = json.load(f)
            mapping = {
                "vincitore":        "VINCITORE",
                "finalista_1":      "FINALISTA 1",
                "finalista_2":      "FINALISTA 2",
                "capocannoniere":   "CAPOCANNONIERE",
                "assistman":        "ASSISTMAN",
                "mvp":              "MVP TORNEO",
                "miglior_portiere": "MIGLIOR PORTIERE",
                "miglior_giovane":  "MIGLIOR GIOVANE U21",
            }
            for key, voce in mapping.items():
                val = rs.get(key, "")
                if val:
                    speciali_esistenti[voce] = val
                    log.info(f"  📋 Speciale da JSON: {voce} = {val}")
    except Exception as e:
        log.warning(f"Errore lettura risultati_speciali.json: {e}")
    # 2. Fallback: leggi dall'Excel esistente per valori non in JSON
    try:
        wb_old = openpyxl.load_workbook(RISULTATI_FILE, data_only=True)
        if "SPECIALI" in wb_old.sheetnames:
            ws_old = wb_old["SPECIALI"]
            for row in ws_old.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    voce = str(row[0]).strip()
                    if voce not in speciali_esistenti:
                        speciali_esistenti[voce] = str(row[1]).strip()
    except Exception:
        pass

    voci = ["VINCITORE", "FINALISTA 1", "FINALISTA 2", "CAPOCANNONIERE",
            "ASSISTMAN", "MVP TORNEO", "MIGLIOR PORTIERE", "MIGLIOR GIOVANE U21"]
    for i, voce in enumerate(voci, 2):
        valore = speciali_esistenti.get(voce, "")
        for col, val in [(1, voce), (2, valore)]:
            c = ws_s.cell(row=i, column=col, value=val)
            c.fill = f_alt[i % 2]
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = bordo()
        ws_s.row_dimensions[i].height = 20

    ws_s.column_dimensions['A'].width = 25
    ws_s.column_dimensions['B'].width = 25
    ws_s.row_dimensions[1].height = 24

    RISULTATI_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(RISULTATI_FILE)
    log.info(f"risultati.xlsx aggiornato: {RISULTATI_FILE}")


def salva_storico_posizioni(punteggi, storico_path: Path) -> None:
    storico_path.parent.mkdir(parents=True, exist_ok=True)
    posizioni = {p.nome_completo: idx + 1 for idx, p in enumerate(punteggi)}
    try:
        with open(storico_path, "w", encoding="utf-8") as f:
            json.dump(posizioni, f, ensure_ascii=False, indent=2)
        log.info(f"Storico posizioni salvato: {storico_path}")
    except Exception as e:
        log.error(f"Errore salvataggio storico: {e}")


def main():
    log.info("=== Avvio aggiornamento automatico ===")

    partite = scarica_risultati()
    gironi  = scarica_gironi()
    scrivi_risultati(partite, gironi)

    from src.config import PRONOSTICI_DIR
    partecipanti = leggi_tutti_pronostici(PRONOSTICI_DIR)
    if not partecipanti:
        log.warning("Nessun partecipante trovato!")
        return

    risultati = leggi_risultati(RISULTATI_FILE)
    punteggi  = calcola_tutti_punteggi(partecipanti, risultati)

    # Carica dati eliminatorie e somma punti al totale
    eliminatorie = carica_eliminatorie()
    if eliminatorie:
        from src.calcolo_eliminatorie import punti_eliminatorie_totali
        for p in punteggi:
            p.pt_eliminatorie = punti_eliminatorie_totali(p.nome_completo, eliminatorie)
        # Riordina includendo i punti eliminatorie nel totale
        punteggi.sort(
            key=lambda x: (
                x.punti_totali + getattr(x, 'pt_eliminatorie', 0),
                x.n_risultati_esatti,
                x.n_marcatori_corretti,
            ),
            reverse=True,
        )
        log.info(f"Punti eliminatorie aggiunti al totale per {len(punteggi)} partecipanti")

    storico_path = Path(__file__).parent.parent / "data" / "storico_posizioni.json"
    salva_storico_posizioni(punteggi, storico_path)
    genera_excel_classifica(punteggi)
    genera_html(
        punteggi,
        n_partite_giocate=len(partite),
        partecipanti=partecipanti,
        risultati=risultati,
        storico_path=storico_path,
        eliminatorie=eliminatorie,
    )

    log.info(f"=== Completato: {len(punteggi)} partecipanti, {len(partite)} partite ===")
    print(f"\n✅ Classifica aggiornata: {len(punteggi)} partecipanti | {len(partite)} partite giocate")


if __name__ == "__main__":
    main()
